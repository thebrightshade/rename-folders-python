import os
import re
import openpyxl

# Get Directories
directories = [name for name in os.listdir('.')]
# print directories
# Load workbook
wb = openpyxl.load_workbook('database.xlsx')
ws = wb['Sheet1']
rows = 20
date_used = {}
for i in xrange(2, rows + 2):
    if ws['A%s' % i].value >= "":
        folder_to_change = [x for x in directories if x.startswith((ws['A%s' % i].value))]
        for y, folder in enumerate(folder_to_change):
            # print folder
            folder_parts = folder.split('_')
            ADB_ID = folder_parts[0]
            # print ADB_ID
            COMPORT = folder_parts[1]
            # print COMPORT
            DATE_PART = folder_parts[2]
            # print DATE_PART
            DATE_SPLIT = re.search('(....)(....)', DATE_PART)
            DATE = DATE_SPLIT.group(2) + DATE_SPLIT.group(1)
            # print DATE
            TIMESTAMP = folder_parts[3]
            # print TIMESTAMP
            os.rename(folder, '{0}_{1}_{2}_{3}_{4}_{5}_{6}_{7}_{8}'.format(ws['B%s' % i].value, ws['C%s' % i].value,
                                                                           ws['G%s' % i].value, DATE,
                                                                           ws['D%s' % i].value, y+1,
                                                                           ws['F%s' % i].value, ws['H%s' % i].value,
                                                                           ws['I%s' % i].value))
